import unittest
import os
import sys
import io
import builtins
import traceback
from datetime import datetime

# Fix encoding cho Windows console
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] Cần cài đặt thư viện openpyxl: pip install openpyxl")
    sys.exit(1)

# 1. Cấu hình đường dẫn
current_directory = os.getcwd()

# Đường dẫn folder chứa test case
users_test_dir = os.path.join(current_directory, 'users')
admin_test_dir = os.path.join(current_directory, 'admin')

# Đường dẫn folder xuất báo cáo
results_dir = os.path.join(current_directory, 'results')
users_report_dir = os.path.join(results_dir, 'Users_test_result')
admin_report_dir = os.path.join(results_dir, 'Admin_test_result')


class ExcelTestResult(unittest.TestResult):
    """Custom TestResult để thu thập kết quả chi tiết theo từng test suite"""
    
    def __init__(self):
        super().__init__()
        self.test_results = {}  # Dict lưu kết quả theo từng test suite
        self.current_suite = None
        self.start_times = {}
        
    def _get_suite_name(self, test):
        """Lấy tên test suite từ test case"""
        # Lấy tên class chứa test
        return test.__class__.__name__
    
    def _get_module_name(self, test):
        """Lấy tên module từ test case"""
        return test.__class__.__module__
    
    def startTest(self, test):
        super().startTest(test)
        self.start_times[test] = datetime.now()
        
        suite_name = self._get_suite_name(test)
        module_name = self._get_module_name(test)
        
        # Tạo key cho suite (module + class name)
        suite_key = f"{module_name}.{suite_name}"
        
        if suite_key not in self.test_results:
            self.test_results[suite_key] = {
                'module': module_name,
                'class': suite_name,
                'tests': []
            }
    
    def stopTest(self, test):
        super().stopTest(test)
        
    def _add_result(self, test, status, message=""):
        suite_name = self._get_suite_name(test)
        module_name = self._get_module_name(test)
        suite_key = f"{module_name}.{suite_name}"
        
        # Tính thời gian thực thi
        start_time = self.start_times.get(test, datetime.now())
        duration = (datetime.now() - start_time).total_seconds()
        
        # Lấy tên method test
        test_method = str(test).split()[0]
        
        # Lấy docstring của test method
        test_doc = test.shortDescription() or ""
        
        self.test_results[suite_key]['tests'].append({
            'test_name': test_method,
            'description': test_doc,
            'status': status,
            'duration': round(duration, 3),
            'message': message
        })
    
    def addSuccess(self, test):
        super().addSuccess(test)
        self._add_result(test, 'PASSED')
    
    def addFailure(self, test, err):
        super().addFailure(test, err)
        error_msg = ''.join(traceback.format_exception(*err))
        self._add_result(test, 'FAILED', error_msg)
    
    def addError(self, test, err):
        super().addError(test, err)
        error_msg = ''.join(traceback.format_exception(*err))
        self._add_result(test, 'ERROR', error_msg)
    
    def addSkip(self, test, reason):
        super().addSkip(test, reason)
        self._add_result(test, 'SKIPPED', reason)


def create_excel_report(test_result, report_path, report_title):
    """Tạo báo cáo Excel từ kết quả test"""
    
    wb = openpyxl.Workbook()
    
    # Xóa sheet mặc định
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Định nghĩa styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    passed_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    passed_font = Font(color='006100', bold=True)
    
    failed_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    failed_font = Font(color='9C0006', bold=True)
    
    error_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    error_font = Font(color='9C5700', bold=True)
    
    skipped_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    skipped_font = Font(color='666666', bold=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet tổng hợp (Summary)
    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(['BÁO CÁO KIỂM THỬ TỰ ĐỘNG'])
    summary_ws.merge_cells('A1:F1')
    summary_ws['A1'].font = Font(bold=True, size=16)
    summary_ws['A1'].alignment = Alignment(horizontal='center')
    
    summary_ws.append([f'Tiêu đề: {report_title}'])
    summary_ws.append([f'Ngày chạy: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
    summary_ws.append([])
    
    # Header cho bảng tổng hợp
    summary_headers = ['STT', 'Test Suite', 'Total', 'Passed', 'Failed', 'Error', 'Skipped', 'Pass Rate']
    summary_ws.append(summary_headers)
    
    header_row = 5
    for col_idx, header in enumerate(summary_headers, 1):
        cell = summary_ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Tổng số liệu
    total_tests = 0
    total_passed = 0
    total_failed = 0
    total_error = 0
    total_skipped = 0
    
    # Tạo sheet cho từng test suite
    suite_idx = 0
    for suite_key, suite_data in test_result.test_results.items():
        suite_idx += 1
        
        # Tạo tên sheet (giới hạn 31 ký tự)
        sheet_name = suite_data['class'][:31]
        
        # Đảm bảo tên sheet không trùng
        original_name = sheet_name
        counter = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{original_name[:28]}_{counter}"
            counter += 1
        
        ws = wb.create_sheet(sheet_name)
        
        # Tiêu đề sheet
        ws.append([f"Test Suite: {suite_data['class']}"])
        ws.merge_cells('A1:F1')
        ws['A1'].font = Font(bold=True, size=14)
        
        ws.append([f"Module: {suite_data['module']}"])
        ws.append([])
        
        # Header
        headers = ['STT', 'Test Name', 'Description', 'Status', 'Duration (s)', 'Message']
        ws.append(headers)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Dữ liệu test
        suite_passed = 0
        suite_failed = 0
        suite_error = 0
        suite_skipped = 0
        
        for idx, test in enumerate(suite_data['tests'], 1):
            row = [
                idx,
                test['test_name'],
                test['description'],
                test['status'],
                test['duration'],
                test['message'][:500] if test['message'] else ''  # Giới hạn message
            ]
            ws.append(row)
            
            row_idx = idx + 4
            
            # Áp dụng style theo status
            status_cell = ws.cell(row=row_idx, column=4)
            if test['status'] == 'PASSED':
                status_cell.fill = passed_fill
                status_cell.font = passed_font
                suite_passed += 1
            elif test['status'] == 'FAILED':
                status_cell.fill = failed_fill
                status_cell.font = failed_font
                suite_failed += 1
            elif test['status'] == 'ERROR':
                status_cell.fill = error_fill
                status_cell.font = error_font
                suite_error += 1
            elif test['status'] == 'SKIPPED':
                status_cell.fill = skipped_fill
                status_cell.font = skipped_font
                suite_skipped += 1
            
            # Áp dụng border cho tất cả cell trong row
            for col_idx in range(1, 7):
                ws.cell(row=row_idx, column=col_idx).border = border
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical='top')
        
        # Điều chỉnh độ rộng cột
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 60
        
        # Thêm dòng tổng kết
        ws.append([])
        ws.append(['', 'TỔNG KẾT:', '', '', '', ''])
        ws.append(['', f'Passed: {suite_passed}', f'Failed: {suite_failed}', f'Error: {suite_error}', f'Skipped: {suite_skipped}', ''])
        
        # Cập nhật tổng số liệu
        suite_total = len(suite_data['tests'])
        total_tests += suite_total
        total_passed += suite_passed
        total_failed += suite_failed
        total_error += suite_error
        total_skipped += suite_skipped
        
        # Thêm vào Summary sheet
        pass_rate = f"{(suite_passed/suite_total*100):.1f}%" if suite_total > 0 else "0%"
        summary_row = [suite_idx, suite_data['class'], suite_total, suite_passed, suite_failed, suite_error, suite_skipped, pass_rate]
        summary_ws.append(summary_row)
        
        row_num = header_row + suite_idx
        for col_idx in range(1, 9):
            cell = summary_ws.cell(row=row_num, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    
    # Thêm dòng tổng vào Summary
    summary_ws.append([])
    total_pass_rate = f"{(total_passed/total_tests*100):.1f}%" if total_tests > 0 else "0%"
    total_row = ['', 'TỔNG CỘNG', total_tests, total_passed, total_failed, total_error, total_skipped, total_pass_rate]
    summary_ws.append(total_row)
    
    total_row_num = summary_ws.max_row
    for col_idx in range(1, 9):
        cell = summary_ws.cell(row=total_row_num, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Điều chỉnh độ rộng cột Summary
    summary_ws.column_dimensions['A'].width = 5
    summary_ws.column_dimensions['B'].width = 40
    summary_ws.column_dimensions['C'].width = 10
    summary_ws.column_dimensions['D'].width = 10
    summary_ws.column_dimensions['E'].width = 10
    summary_ws.column_dimensions['F'].width = 10
    summary_ws.column_dimensions['G'].width = 10
    summary_ws.column_dimensions['H'].width = 12
    
    # Di chuyển Summary lên đầu
    wb.move_sheet(summary_ws, offset=-len(wb.sheetnames)+1)
    
    # Lưu file
    wb.save(report_path)
    print(f"[REPORT] Đã tạo báo cáo Excel: {report_path}")
    
    return total_tests, total_passed, total_failed, total_error, total_skipped


def run_users_tests_excel():
    """Chạy tất cả test cases cho luồng Users và xuất Excel"""
    print("\n" + "="*60)
    print("[TEST] BẮT ĐẦU CHẠY TEST LUỒNG USERS")
    print("="*60 + "\n")
    
    # Tạo folder kết quả nếu chưa có
    if not os.path.exists(users_report_dir):
        os.makedirs(users_report_dir)
    
    # Tự động tìm tất cả các Test Case trong folder users
    loader = unittest.TestLoader()
    suite = loader.discover(start_dir=users_test_dir, pattern='test_suite_*.py')
    
    # Chạy test với custom result
    result = ExcelTestResult()
    suite.run(result)
    
    # Tạo báo cáo Excel
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    report_path = os.path.join(users_report_dir, f"Users_Module_Report_{timestamp}.xlsx")
    
    stats = create_excel_report(
        result, 
        report_path, 
        "Báo Cáo Kiểm Thử Tự Động - Luồng Người Dùng"
    )
    
    return result, report_path, stats


def run_admin_tests_excel():
    """Chạy tất cả test cases cho luồng Admin và xuất Excel"""
    print("\n" + "="*60)
    print("[TEST] BẮT ĐẦU CHẠY TEST LUỒNG ADMIN")
    print("="*60 + "\n")
    
    # Tạo folder kết quả nếu chưa có
    if not os.path.exists(admin_report_dir):
        os.makedirs(admin_report_dir)
    
    # Tự động tìm tất cả các Test Case trong folder admin
    loader = unittest.TestLoader()
    suite = loader.discover(start_dir=admin_test_dir, pattern='test_suite_*.py')
    
    # Chạy test với custom result
    result = ExcelTestResult()
    suite.run(result)
    
    # Tạo báo cáo Excel
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    report_path = os.path.join(admin_report_dir, f"Admin_Module_Report_{timestamp}.xlsx")
    
    stats = create_excel_report(
        result, 
        report_path, 
        "Báo Cáo Kiểm Thử Tự Động - Luồng Quản Trị"
    )
    
    return result, report_path, stats


def print_summary(users_data, admin_data):
    """In tổng kết kết quả test"""
    print("\n" + "="*60)
    print("[SUMMARY] TỔNG KẾT KẾT QUẢ KIỂM THỬ")
    print("="*60)
    
    # Thống kê Users
    if users_data:
        result, report_path, stats = users_data
        total, passed, failed, error, skipped = stats
        
        print(f"\n[USERS] LUỒNG USERS:")
        print(f"   [PASS]    Passed:  {passed}")
        print(f"   [FAIL]    Failed:  {failed}")
        print(f"   [ERROR]   Errors:  {error}")
        print(f"   [SKIP]    Skipped: {skipped}")
        print(f"   [TOTAL]   Total:   {total}")
        print(f"   [REPORT]  Excel:   {report_path}")
    
    # Thống kê Admin
    if admin_data:
        result, report_path, stats = admin_data
        total, passed, failed, error, skipped = stats
        
        print(f"\n[ADMIN] LUỒNG ADMIN:")
        print(f"   [PASS]    Passed:  {passed}")
        print(f"   [FAIL]    Failed:  {failed}")
        print(f"   [ERROR]   Errors:  {error}")
        print(f"   [SKIP]    Skipped: {skipped}")
        print(f"   [TOTAL]   Total:   {total}")
        print(f"   [REPORT]  Excel:   {report_path}")
    
    print("\n" + "="*60)
    print("[DONE] HOÀN TẤT KIỂM THỬ!")
    print("="*60 + "\n")


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Chạy kiểm thử tự động cho Web Thể Thao - Xuất Excel')
    parser.add_argument('--users', action='store_true', help='Chỉ chạy test luồng Users')
    parser.add_argument('--admin', action='store_true', help='Chỉ chạy test luồng Admin')
    parser.add_argument('--all', action='store_true', help='Chạy tất cả test (mặc định)')
    
    args = parser.parse_args()
    
    users_data = None
    admin_data = None
    
    # Nếu không có tham số nào, chạy tất cả
    if not args.users and not args.admin:
        args.all = True
    
    # Chạy test theo tham số
    if args.users or args.all:
        users_data = run_users_tests_excel()
    
    if args.admin or args.all:
        admin_data = run_admin_tests_excel()
    
    # In tổng kết
    print_summary(users_data, admin_data)
    
    # Mở báo cáo Excel
    if args.users or args.all:
        if users_data:
            _, report_path, _ = users_data
            os.startfile(report_path)
    
    if args.admin or args.all:
        if admin_data:
            _, report_path, _ = admin_data
            os.startfile(report_path)
